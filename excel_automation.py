from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.styles import Alignment
import os
import re

class EasyExcelInput:
    def __init__(self):
        self.wb = None
        self.ws = None
        self.current_row = 5
        self.start_col = 'F'
        self.col_num = 6
        self.file_path = None
    
    def run(self):
        print("=" * 60)
        print("     엑셀 측정 데이터 입력 시스템")
        print("=" * 60)
        
        # 1. 파일 선택
        self.select_file()
        
        # 2. 시트 선택 추가
        self.select_sheet()
        
        # 3. 초기 설정
        self.initialize()
        
        # 4. 메인 루프
        while True:
            self.show_menu()
            choice = input("\n선택하세요: ").strip()
            
            if choice == '1':
                self.input_simple()
            elif choice == '2':
                self.input_position()
            elif choice == '3':
                self.input_reference()
            elif choice == '4':
                self.input_mmc()
            elif choice == '5':
                self.input_batch()
            elif choice == '6':
                self.save_and_exit()
                break
            elif choice == '7':
                self.save_file()
            elif choice == '8':
                self.show_position()
            elif choice == '9':
                self.change_position()
            elif choice == 's':
                self.change_sheet()
            elif choice == '0':
                print("\n종료합니다.")
                break
            else:
                print("\n❌ 잘못된 선택입니다!")
    
    def clean_path(self, path):
        """파일 경로 정리"""
        path = re.sub(r'^&\s*', '', path)
        path = path.strip()
        path = path.strip('"').strip("'")
        path = path.replace('\\', '/')
        return path
    
    def select_file(self):
        print("\n📁 파일 경로를 입력하세요:")
        print("   (파일을 드래그 앤 드롭 하거나 전체 경로를 입력)")
        print("   💡 팁: Windows에서 Shift + 우클릭 → '경로 복사'를 사용하세요")
        print()
        
        while True:
            path = input("파일 경로: ").strip()
            
            cleaned_path = self.clean_path(path)
            
            if os.path.exists(cleaned_path):
                if cleaned_path.lower().endswith(('.xlsx', '.xlsm')):
                    self.file_path = cleaned_path
                    print(f"✓ 파일 선택 완료: {os.path.basename(cleaned_path)}\n")
                    
                    # 파일 열기 (시트 목록 확인용)
                    try:
                        self.wb = load_workbook(cleaned_path)
                        print(f"✓ 파일 로드 완료 (시트 {len(self.wb.sheetnames)}개 발견)")
                    except Exception as e:
                        print(f"❌ 파일 열기 실패: {e}")
                        print("   파일이 손상되었거나 다른 프로그램에서 사용 중일 수 있습니다.\n")
                        continue
                    
                    break
                else:
                    print("❌ Excel 파일(.xlsx, .xlsm)만 사용 가능합니다.\n")
            else:
                alt_path = cleaned_path.replace('/', '\\')
                if os.path.exists(alt_path):
                    self.file_path = alt_path
                    print(f"✓ 파일 선택 완료: {os.path.basename(alt_path)}\n")
                    
                    try:
                        self.wb = load_workbook(alt_path)
                        print(f"✓ 파일 로드 완료 (시트 {len(self.wb.sheetnames)}개 발견)")
                    except Exception as e:
                        print(f"❌ 파일 열기 실패: {e}\n")
                        continue
                    
                    break
                
                print("❌ 파일을 찾을 수 없습니다.")
                print("   다음을 확인해주세요:")
                print("   1. 파일이 실제로 존재하는지")
                print("   2. 경로에 특수문자가 없는지")
                print("   3. 파일이 다른 프로그램에서 열려있지 않은지")
                print()
                
                retry = input("다시 시도하시겠습니까? (y/n): ").strip().lower()
                if retry != 'y':
                    print("프로그램을 종료합니다.")
                    exit(0)
    
    def select_sheet(self):
        """시트 선택"""
        print("\n" + "=" * 60)
        print("📊 시트 선택")
        print("=" * 60)
        
        sheet_names = self.wb.sheetnames
        
        if len(sheet_names) == 1:
            self.ws = self.wb[sheet_names[0]]
            print(f"✓ 시트가 1개만 있어 자동 선택: '{sheet_names[0]}'")
            return
        
        # 시트 목록 표시
        print("\n사용 가능한 시트 목록:")
        print("-" * 60)
        for i, name in enumerate(sheet_names, 1):
            # 활성 시트 표시
            is_active = "⭐ (현재 활성)" if self.wb.active == self.wb[name] else ""
            print(f"{i}. {name} {is_active}")
        print("-" * 60)
        
        # 시트 선택
        while True:
            choice = input(f"\n시트 번호를 선택하세요 (1-{len(sheet_names)}, Enter=활성시트): ").strip()
            
            # Enter만 누르면 활성 시트 사용
            if not choice:
                self.ws = self.wb.active
                print(f"✓ 활성 시트 선택: '{self.ws.title}'")
                break
            
            # 숫자로 선택
            try:
                idx = int(choice) - 1
                if 0 <= idx < len(sheet_names):
                    self.ws = self.wb[sheet_names[idx]]
                    print(f"✓ 시트 선택 완료: '{self.ws.title}'")
                    break
                else:
                    print(f"❌ 1에서 {len(sheet_names)} 사이의 숫자를 입력해주세요.")
            except ValueError:
                # 시트 이름으로 직접 선택
                if choice in sheet_names:
                    self.ws = self.wb[choice]
                    print(f"✓ 시트 선택 완료: '{self.ws.title}'")
                    break
                else:
                    print(f"❌ '{choice}' 시트를 찾을 수 없습니다. 숫자나 정확한 시트 이름을 입력해주세요.")
    
    def change_sheet(self):
        """작업 중 시트 변경"""
        print("\n" + "=" * 60)
        print("📊 시트 변경")
        print("=" * 60)
        
        sheet_names = self.wb.sheetnames
        
        print("\n사용 가능한 시트 목록:")
        print("-" * 60)
        for i, name in enumerate(sheet_names, 1):
            current = "✓ (현재 작업 중)" if self.ws.title == name else ""
            print(f"{i}. {name} {current}")
        print("-" * 60)
        
        while True:
            choice = input(f"\n새 시트 번호를 선택하세요 (1-{len(sheet_names)}, 0=취소): ").strip()
            
            if choice == '0':
                print("취소되었습니다.")
                return
            
            try:
                idx = int(choice) - 1
                if 0 <= idx < len(sheet_names):
                    old_sheet = self.ws.title
                    self.ws = self.wb[sheet_names[idx]]
                    print(f"✓ 시트 변경: '{old_sheet}' → '{self.ws.title}'")
                    
                    # 위치 초기화 여부 확인
                    reset = input(f"\n현재 위치({self.start_col}{self.current_row})를 초기화하시겠습니까? (y/n): ").strip().lower()
                    if reset == 'y':
                        self.current_row = 5
                        print(f"✓ 위치 초기화: {self.start_col}{self.current_row}")
                    
                    break
                else:
                    print(f"❌ 1에서 {len(sheet_names)} 사이의 숫자를 입력해주세요.")
            except ValueError:
                if choice in sheet_names:
                    old_sheet = self.ws.title
                    self.ws = self.wb[choice]
                    print(f"✓ 시트 변경: '{old_sheet}' → '{self.ws.title}'")
                    
                    reset = input(f"\n현재 위치({self.start_col}{self.current_row})를 초기화하시겠습니까? (y/n): ").strip().lower()
                    if reset == 'y':
                        self.current_row = 5
                        print(f"✓ 위치 초기화: {self.start_col}{self.current_row}")
                    
                    break
                else:
                    print(f"❌ '{choice}' 시트를 찾을 수 없습니다.")
    
    def initialize(self):
        print("\n⚙️ 초기 설정")
        
        # 시작 열
        col = input(f"시작 열 (기본값: F, Enter로 건너뛰기): ").strip().upper()
        if col:
            self.start_col = col
        
        # 시작 행
        row = input(f"시작 행 (기본값: 5, Enter로 건너뛰기): ").strip()
        if row:
            self.current_row = max(int(row), 5)
        
        # 컬럼 번호 계산
        try:
            self.col_num = column_index_from_string(self.start_col)
            
            print(f"\n✓ 초기화 완료!")
            print(f"  - 파일: {os.path.basename(self.file_path)}")
            print(f"  - 시트: {self.ws.title}")
            print(f"  - 시작 위치: {self.start_col}{self.current_row}")
            
        except Exception as e:
            print(f"\n❌ 초기화 실패: {e}")
            exit(1)
    
    def show_menu(self):
        print("\n" + "=" * 60)
        print(f"  파일: {os.path.basename(self.file_path)}")
        print(f"  시트: {self.ws.title}")
        print(f"  현재 위치: {self.start_col}{self.current_row}")
        print("=" * 60)
        print("1. 단순 측정값 입력")
        print("2. 위치도 값 입력")
        print("3. 참고 값 입력")
        print("4. MMC 공차 입력")
        print("5. 일괄 입력 (자동 감지)")
        print("-" * 60)
        print("6. 저장 후 종료")
        print("7. 저장 (계속)")
        print("8. 현재 위치 확인")
        print("9. 위치 변경")
        print("s. 시트 변경")
        print("0. 종료 (저장 안함)")
        print("=" * 60)
    
    def get_input(self, prompt, default="", required=True):
        """입력 받기 헬퍼 함수"""
        while True:
            value = input(f"{prompt}: ").strip()
            if value:
                return value
            elif not required or default:
                return default
            else:
                print("❌ 필수 입력 항목입니다. 다시 입력해주세요.")
    
    def input_simple(self):
        print("\n📝 단순 측정값 입력")
        print("=" * 40)
        
        try:
            item = self.get_input("항목 번호", required=True)
            base = float(self.get_input("기준값", required=True))
            plus = float(self.get_input("+공차", required=True))
            minus = float(self.get_input("-공차", required=True))
            ref = self.get_input("REF (선택, Enter로 건너뛰기)", required=False)
            
            upper = base + plus
            lower = base - minus
            
            self.ws.cell(self.current_row, self.col_num, item)
            self.ws.cell(self.current_row, self.col_num + 1, base)
            self.ws.cell(self.current_row, self.col_num + 2, upper)
            self.ws.cell(self.current_row, self.col_num + 3, lower)
            self.ws.cell(self.current_row, self.col_num + 4, upper)
            self.ws.cell(self.current_row, self.col_num + 5, lower)
            self.ws.cell(self.current_row, self.col_num + 6, ref)
            
            print(f"✓ {self.ws.title} 시트의 {self.current_row}행에 추가되었습니다!")
            self.current_row += 1
            
        except ValueError:
            print("❌ 숫자 입력 오류! 다시 시도해주세요.")
        except Exception as e:
            print(f"❌ 오류 발생: {e}")
    
    def input_position(self):
        print("\n📝 위치도 값 입력")
        print("=" * 40)
        
        try:
            item = self.get_input("항목 번호", required=True)
            base = self.get_input("기준값 (예: Ø4.25)", required=True)
            upper = float(self.get_input("상한값", required=True))
            lower = float(self.get_input("하한값", required=True))
            rows = int(self.get_input("행 개수", default="4"))
            ref = self.get_input("REF (선택, Enter로 건너뛰기)", required=False)
            
            for i in range(rows):
                row = self.current_row + i
                self.ws.cell(row, self.col_num, f"{item}-{i+1}")
                self.ws.cell(row, self.col_num + 1, base)
                self.ws.cell(row, self.col_num + 2, upper)
                self.ws.cell(row, self.col_num + 3, lower)
                self.ws.cell(row, self.col_num + 4, upper)
                self.ws.cell(row, self.col_num + 5, lower)
                self.ws.cell(row, self.col_num + 6, ref)
            
            print(f"✓ {self.ws.title} 시트의 {self.current_row}행부터 {rows}개 행 추가되었습니다!")
            self.current_row += rows
            
        except ValueError:
            print("❌ 숫자 입력 오류! 다시 시도해주세요.")
        except Exception as e:
            print(f"❌ 오류 발생: {e}")
    
    def input_reference(self):
        print("\n📝 참고 값 입력")
        print("=" * 40)
        
        try:
            item = self.get_input("항목 번호", required=True)
            base = self.get_input("기준값", required=True)
            ref = self.get_input("REF", default="참고")
            
            self.ws.cell(self.current_row, self.col_num, item)
            self.ws.cell(self.current_row, self.col_num + 1, base)
            for i in range(2, 7):
                self.ws.cell(self.current_row, self.col_num + i, '-')
            self.ws.cell(self.current_row, self.col_num + 6, ref)
            
            print(f"✓ {self.ws.title} 시트의 {self.current_row}행에 추가되었습니다!")
            self.current_row += 1
            
        except Exception as e:
            print(f"❌ 오류 발생: {e}")
    
    def input_mmc(self):
        print("\n📝 MMC 공차 입력")
        print("=" * 40)
        
        try:
            item = self.get_input("항목 번호", required=True)
            base = float(self.get_input("기준값", required=True))
            mmc = float(self.get_input("MMC 허용공차", required=True))
            upper = float(self.get_input("상한값", required=True))
            lower = float(self.get_input("하한값", required=True))
            ref = self.get_input("REF (선택, Enter로 건너뛰기)", required=False)
            
            # 1행
            self.ws.cell(self.current_row, self.col_num, f"{item}-1")
            self.ws.cell(self.current_row, self.col_num + 1, base)
            self.ws.cell(self.current_row, self.col_num + 2, upper)
            self.ws.cell(self.current_row, self.col_num + 3, lower)
            self.ws.cell(self.current_row, self.col_num + 4, upper)
            self.ws.cell(self.current_row, self.col_num + 5, lower)
            self.ws.cell(self.current_row, self.col_num + 6, ref)
            
            # 2행
            self.ws.cell(self.current_row + 1, self.col_num, f"{item}-2")
            self.ws.cell(self.current_row + 1, self.col_num + 1, f"MMC: {mmc}")
            for i in range(2, 7):
                self.ws.cell(self.current_row + 1, self.col_num + i, '-')
            self.ws.cell(self.current_row + 1, self.col_num + 6, ref)
            
            # 3-4행
            for offset in [2, 3]:
                self.ws.cell(self.current_row + offset, self.col_num, f"{item}-{offset+1}")
                self.ws.cell(self.current_row + offset, self.col_num + 1, f"계산{offset-1}")
                self.ws.cell(self.current_row + offset, self.col_num + 6, ref)
            
            print(f"✓ {self.ws.title} 시트의 {self.current_row}행부터 4개 행 추가되었습니다!")
            self.current_row += 4
            
        except ValueError:
            print("❌ 숫자 입력 오류! 다시 시도해주세요.")
        except Exception as e:
            print(f"❌ 오류 발생: {e}")
    
    def input_batch(self):
        """통합 일괄 입력 - 자동 유형 감지"""
        print("\n📝 일괄 입력")
        print("=" * 60)
        print("입력 형식:")
        print()
        print("1. 단순 측정값 (5개 항목):")
        print("   항목번호, 행개수, 기준값, 상한공차, 하한공차")
        print("   예: 51, 1, 7.0, 0.15, 0.15")
        print("   → 7.0, 0.15, -0.15, 6.85, 7.15")
        print()
        print("2. 위치도 값 (6개 항목, 기준값에 Ø 포함):")
        print("   항목번호, 행개수, 기준값, 상한공차, 하한공차, [REF]")
        print("   예: 55, 4, Ø4.25, 0.15, 0.15")
        print()
        print("3. 참고 값 (기준값이 괄호로 싸여있음):")
        print("   항목번호, 행개수, (기준값), [상한공차], [하한공차], [REF]")
        print("   예: 60, 1, (1.2) 또는 61, 3, (7.0), 0.15, 0.15, 참고")
        print()
        print("4. MMC 공차 (MMC공차에 'm' 포함):")
        print("   항목번호, 세트개수, MMC공차, [MAX값]")
        print("   예: 70, 10, 0.2m 또는 70, 10, 0.2m, 0.5")
        print("   → 각 세트: 기준값행, MAX값행(MMC공차), 측정값행(빈칸)")
        print()
        print(f"현재 위치 {self.start_col}{self.current_row}부터 입력됩니다.")
        print("\n여러 줄 입력 후 빈 줄로 완료:")

        lines = []
        while True:
            line = input().strip()
            if not line:
                break
            lines.append(line)

        if not lines:
            print("❌ 입력된 데이터가 없습니다.")
            return

        count = 0
        start_row = self.current_row

        for line in lines:
            parts = [p.strip() for p in line.split(',')]
            if len(parts) < 2:
                print(f"⚠ 형식 오류 (최소 2개 항목 필요): {line}")
                continue

            try:
                # 자동 유형 감지
                data_type = self._detect_data_type(parts, line)

                if data_type == 'simple':
                    self._batch_simple(parts, line)
                elif data_type == 'position':
                    self._batch_position(parts, line)
                elif data_type == 'reference':
                    self._batch_reference(parts, line)
                elif data_type == 'mmc':
                    self._batch_mmc(parts, line)
                else:
                    print(f"⚠ 유형을 감지할 수 없음: {line}")
                    continue

                count += 1

            except Exception as e:
                print(f"  ⚠ 오류: {line} - {e}")

        print(f"\n✓ 총 {count}개 항목이 {start_row}행부터 입력되었습니다!")
        print(f"  다음 입력 위치: {self.start_col}{self.current_row}")

    def _detect_data_type(self, parts, line):
        """데이터 유형 자동 감지"""
        line_lower = line.lower()

        # 1. 참고 값 감지 - 기준값이 괄호로 싸여있거나 REF/참고 키워드 포함
        if len(parts) >= 3:
            # 괄호로 싸여있는 경우: 50, 1, (1.2) 또는 50, 3, (1.2), 0.15, 0.15
            if '(' in parts[2] and ')' in parts[2]:
                return 'reference'
            # REF/참고 키워드가 있는 경우
            if 'ref' in line_lower or '참고' in line_lower:
                return 'reference'

        # 2. MMC 공차 감지 ('mmc' 또는 'm' 문자 포함)
        # parts[2]에 'm'이 있으면 MMC (예: 70, 10, 0.2m)
        if 'mmc' in line_lower or (len(parts) >= 3 and 'm' in parts[2].lower() and not 'mm' in parts[2].lower()):
            return 'mmc'

        # 3. 위치도 값 감지 (기준값에 Ø 포함 또는 6개 항목)
        if len(parts) >= 5:
            # 3번째 항목(기준값)에 Ø가 있으면 위치도
            if 'ø' in parts[2].lower() or 'Ø' in parts[2]:
                return 'position'
            # 6개 항목이면 위치도로 판단
            if len(parts) == 6:
                try:
                    # 4, 5번째가 숫자면 위치도 (상한, 하한)
                    float(parts[3])
                    float(parts[4])
                    return 'position'
                except:
                    pass

        # 4. 단순 측정값 (5개 항목, 기본값)
        if len(parts) >= 5:
            return 'simple'

        return None

    def _unmerge_range(self, start_row, num_rows):
        """지정된 범위의 모든 병합 해제"""
        # 해당 범위의 모든 병합된 셀 찾기
        merged_ranges_to_unmerge = []
        for merged_range in list(self.ws.merged_cells.ranges):
            # 병합 범위가 우리가 쓰려는 영역과 겹치는지 확인
            if (merged_range.min_row <= start_row + num_rows - 1 and
                merged_range.max_row >= start_row and
                merged_range.min_col <= self.col_num + 6 and
                merged_range.max_col >= self.col_num):
                merged_ranges_to_unmerge.append(merged_range)

        # 찾은 병합 해제
        for merged_range in merged_ranges_to_unmerge:
            self.ws.unmerge_cells(str(merged_range))

    def _batch_simple(self, parts, line):
        """단순 측정값 일괄 입력"""
        if len(parts) < 5:
            raise ValueError("형식: 항목번호, 행개수, 기준값, 상한공차, 하한공차")

        item_no = parts[0]
        rows = int(parts[1])
        base = float(parts[2].replace('Ø', '').replace('ø', ''))
        upper_tol = float(parts[3])  # 상한공차 (양수)
        lower_tol = float(parts[4])  # 하한공차
        ref = parts[5] if len(parts) > 5 else ""

        # 하한공차가 양수로 입력되면 자동으로 마이너스 붙이기
        if lower_tol > 0:
            lower_tol = -lower_tol

        # 계산값
        lower_calc = base + lower_tol  # 하한계산값 (기준 + 하한공차)
        upper_calc = base + upper_tol  # 상한계산값 (기준 + 상한공차)

        start_row = self.current_row

        # 기존 병합 해제
        self._unmerge_range(start_row, rows)

        for i in range(rows):
            row = self.current_row + i
            # 항목번호는 첫 행에만 입력 (병합할 것이므로)
            if i == 0:
                self.ws.cell(row, self.col_num, item_no)
            self.ws.cell(row, self.col_num + 1, base)
            self.ws.cell(row, self.col_num + 2, upper_tol)
            self.ws.cell(row, self.col_num + 3, lower_tol)
            self.ws.cell(row, self.col_num + 4, lower_calc)
            self.ws.cell(row, self.col_num + 5, upper_calc)
            self.ws.cell(row, self.col_num + 6, ref)

        # 여러 행이면 항목번호 셀 병합
        if rows > 1:
            # 새로 병합
            self.ws.merge_cells(start_row=start_row, start_column=self.col_num,
                              end_row=start_row + rows - 1, end_column=self.col_num)
            # 병합된 셀 중앙 정렬
            self.ws.cell(start_row, self.col_num).alignment = Alignment(horizontal='center', vertical='center')

        self.current_row += rows
        print(f"  ✓ [단순] 항목 {item_no}: {rows}개 행")

    def _batch_position(self, parts, line):
        """위치도 값 일괄 입력"""
        if len(parts) < 5:
            raise ValueError("형식: 항목번호, 행개수, 기준값, 상한공차, 하한공차")

        item_no = parts[0]
        rows = int(parts[1])
        base_str = parts[2]  # 문자열 그대로 (Ø4.25)
        # Ø 제거하고 숫자 추출
        base = float(base_str.replace('Ø', '').replace('ø', ''))
        upper_tol = float(parts[3])  # 상한공차
        lower_tol = float(parts[4])  # 하한공차
        ref = parts[5] if len(parts) > 5 else ""

        # 하한공차가 양수로 입력되면 자동으로 마이너스 붙이기
        if lower_tol > 0:
            lower_tol = -lower_tol

        # 계산값
        lower_calc = base + lower_tol  # 하한계산값
        upper_calc = base + upper_tol  # 상한계산값

        start_row = self.current_row

        # 기존 병합 해제
        self._unmerge_range(start_row, rows)

        for i in range(rows):
            row = self.current_row + i
            # 항목번호는 첫 행에만 입력 (병합할 것이므로)
            if i == 0:
                self.ws.cell(row, self.col_num, item_no)
            self.ws.cell(row, self.col_num + 1, base_str)  # Ø 포함된 문자열
            self.ws.cell(row, self.col_num + 2, upper_tol)
            self.ws.cell(row, self.col_num + 3, lower_tol)
            self.ws.cell(row, self.col_num + 4, lower_calc)
            self.ws.cell(row, self.col_num + 5, upper_calc)
            self.ws.cell(row, self.col_num + 6, ref)

        # 여러 행이면 항목번호 셀 병합
        if rows > 1:
            # 새로 병합
            self.ws.merge_cells(start_row=start_row, start_column=self.col_num,
                              end_row=start_row + rows - 1, end_column=self.col_num)
            # 병합된 셀 중앙 정렬
            self.ws.cell(start_row, self.col_num).alignment = Alignment(horizontal='center', vertical='center')

        self.current_row += rows
        print(f"  ✓ [위치도] 항목 {item_no}: {rows}개 행")

    def _batch_reference(self, parts, line):
        """참고 값 일괄 입력 - 괄호로 감지, 상한/하한 선택적"""
        if len(parts) < 3:
            raise ValueError("형식: 항목번호, 행개수, (기준값), [상한공차], [하한공차], [REF]")

        item_no = parts[0]
        rows = int(parts[1])

        # 괄호는 유지, 공백만 제거
        base_str_with_paren = parts[2].strip()
        # 계산용으로만 괄호 제거
        base_str_calc = parts[2].replace('(', '').replace(')', '').strip()

        # 상한/하한이 있는지 확인
        has_tolerances = len(parts) >= 5 and parts[3] and parts[4]

        start_row = self.current_row

        # 기존 병합 해제
        self._unmerge_range(start_row, rows)

        for i in range(rows):
            row = self.current_row + i

            # 항목번호는 첫 행에만
            if i == 0:
                self.ws.cell(row, self.col_num, item_no)

            self.ws.cell(row, self.col_num + 1, base_str_with_paren)  # 괄호 포함

            if has_tolerances:
                try:
                    base = float(base_str_calc)  # 괄호 제거한 값으로 계산
                    upper_tol = float(parts[3])
                    lower_tol = float(parts[4])

                    # 하한공차가 양수로 입력되면 자동으로 마이너스 붙이기
                    if lower_tol > 0:
                        lower_tol = -lower_tol

                    # 계산값
                    lower_calc = base + lower_tol
                    upper_calc = base + upper_tol

                    self.ws.cell(row, self.col_num + 2, upper_tol)
                    self.ws.cell(row, self.col_num + 3, lower_tol)
                    self.ws.cell(row, self.col_num + 4, lower_calc)
                    self.ws.cell(row, self.col_num + 5, upper_calc)
                except ValueError:
                    # 숫자 변환 실패하면 '-'로 표시
                    for j in range(2, 6):
                        self.ws.cell(row, self.col_num + j, '-')
            else:
                # 상한/하한 없으면 '-'로 표시
                for j in range(2, 6):
                    self.ws.cell(row, self.col_num + j, '-')

            # REF 설정
            if has_tolerances and len(parts) > 5:
                ref = parts[5]
            elif not has_tolerances and len(parts) > 3:
                ref = parts[3]
            else:
                ref = "참고"

            self.ws.cell(row, self.col_num + 6, ref)

        # 여러 행이면 항목번호 셀 병합
        if rows > 1:
            self.ws.merge_cells(start_row=start_row, start_column=self.col_num,
                              end_row=start_row + rows - 1, end_column=self.col_num)
            self.ws.cell(start_row, self.col_num).alignment = Alignment(horizontal='center', vertical='center')

        self.current_row += rows
        print(f"  ✓ [참고] 항목 {item_no}: {rows}개 행")

    def _batch_mmc(self, parts, line):
        """MMC 공차 일괄 입력 - 새로운 형식"""
        # 형식: 항목번호, 세트개수, MMC공차, [MAX값]
        if len(parts) < 3:
            raise ValueError("형식: 항목번호, 세트개수, MMC공차, [MAX값]")

        item_no = parts[0]

        # 세트 개수
        num_sets = int(parts[1])

        # MMC 공차 추출 (0.35m 형식)
        mmc_str = parts[2].lower().replace('mmc', '').replace('(', '').replace(')', '').replace('m', '').strip()
        mmc_tol = float(mmc_str)

        # MAX값 (선택적)
        max_val = parts[3] if len(parts) > 3 and parts[3] else ""

        # REF (선택적)
        ref = parts[4] if len(parts) > 4 else ""

        # 총 행 수 = 세트 개수 * 3
        total_rows = num_sets * 3
        start_row = self.current_row

        # 기존 병합 해제
        self._unmerge_range(start_row, total_rows)

        # 각 세트마다 3개 행 생성
        for set_idx in range(num_sets):
            base_row = self.current_row + (set_idx * 3)

            # 1행: MMC 기준값 행
            if set_idx == 0:
                self.ws.cell(base_row, self.col_num, item_no)  # 첫 세트만 항목번호
            self.ws.cell(base_row, self.col_num + 1, f"{mmc_tol}ⓜ")  # 기준값: 0.2ⓜ
            self.ws.cell(base_row, self.col_num + 2, 0)  # 상한공차 0
            self.ws.cell(base_row, self.col_num + 3, mmc_tol)  # 하한공차 (양수)
            self.ws.cell(base_row, self.col_num + 4, 0)  # 하한계산값
            self.ws.cell(base_row, self.col_num + 5, mmc_tol)  # 상한계산값
            self.ws.cell(base_row, self.col_num + 6, ref)

            # 2행: MAX값 행, REF열에 "MMC 공차"
            if max_val:
                try:
                    self.ws.cell(base_row + 1, self.col_num + 1, float(max_val))
                except:
                    self.ws.cell(base_row + 1, self.col_num + 1, max_val)
            for i in range(2, 6):
                self.ws.cell(base_row + 1, self.col_num + i, '-')
            self.ws.cell(base_row + 1, self.col_num + 6, "MMC 공차")

            # 3행: 측정값 입력 빈 칸
            # 기준값 열만 비우고 나머지는 '-'
            for i in range(2, 6):
                self.ws.cell(base_row + 2, self.col_num + i, '-')
            self.ws.cell(base_row + 2, self.col_num + 6, ref)

        # 항목번호 셀 병합 (전체 행)
        self.ws.merge_cells(start_row=start_row, start_column=self.col_num,
                          end_row=start_row + total_rows - 1, end_column=self.col_num)
        # 병합된 셀 중앙 정렬
        self.ws.cell(start_row, self.col_num).alignment = Alignment(horizontal='center', vertical='center')

        self.current_row += total_rows
        print(f"  ✓ [MMC] 항목 {item_no}: {num_sets}세트 ({total_rows}개 행)")
    
    def save_file(self):
        try:
            print("\n저장 중...")
            self.wb.save(self.file_path)
            print(f"✓ 저장 완료: {os.path.basename(self.file_path)}")
            print(f"  경로: {self.file_path}")
        except Exception as e:
            print(f"❌ 저장 실패: {e}")
            print("  파일이 다른 프로그램에서 열려있다면 닫아주세요.")
    
    def save_and_exit(self):
        self.save_file()
        print("\n프로그램을 종료합니다.")
    
    def show_position(self):
        print(f"\n📍 현재 상태:")
        print(f"  - 시트: {self.ws.title}")
        print(f"  - 위치: {self.start_col}{self.current_row}")
    
    def change_position(self):
        try:
            new_row = int(input(f"\n새로운 행 번호 (현재: {self.current_row}): "))
            self.current_row = max(new_row, 5)
            print(f"✓ 위치 변경: {self.start_col}{self.current_row}")
        except ValueError:
            print("❌ 올바른 숫자를 입력해주세요.")

if __name__ == "__main__":
    try:
        app = EasyExcelInput()
        app.run()
    except KeyboardInterrupt:
        print("\n\n프로그램이 사용자에 의해 중단되었습니다.")
    except Exception as e:
        print(f"\n❌ 예상치 못한 오류: {e}")
        input("\nEnter를 눌러 종료...")