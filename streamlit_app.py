import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import column_index_from_string
import io
import re

class StreamlitExcelInput:
    def __init__(self, wb, sheet_name, start_col, start_row):
        self.wb = wb
        self.ws = wb[sheet_name]
        self.current_row = start_row
        self.start_col = start_col
        self.col_num = column_index_from_string(start_col)

    def _unmerge_range(self, start_row, num_rows):
        """지정된 범위의 모든 병합 해제"""
        merged_ranges_to_unmerge = []
        for merged_range in list(self.ws.merged_cells.ranges):
            if (merged_range.min_row <= start_row + num_rows - 1 and
                merged_range.max_row >= start_row and
                merged_range.min_col <= self.col_num + 6 and
                merged_range.max_col >= self.col_num):
                merged_ranges_to_unmerge.append(merged_range)

        for merged_range in merged_ranges_to_unmerge:
            self.ws.unmerge_cells(str(merged_range))

    def _detect_data_type(self, parts, line):
        """데이터 유형 자동 감지"""
        line_lower = line.lower()

        # 1. 참고 값 감지
        if len(parts) >= 3:
            if '(' in parts[2] and ')' in parts[2]:
                return 'reference'
            if 'ref' in line_lower or '참고' in line_lower:
                return 'reference'

        # 2. MMC 공차 감지
        if 'mmc' in line_lower or (len(parts) >= 3 and 'm' in parts[2].lower() and not 'mm' in parts[2].lower()):
            return 'mmc'

        # 3. 위치도 값 감지
        if len(parts) >= 5:
            if 'ø' in parts[2].lower() or 'Ø' in parts[2]:
                return 'position'
            if len(parts) == 6:
                try:
                    float(parts[3])
                    float(parts[4])
                    return 'position'
                except:
                    pass

        # 4. 단순 측정값
        if len(parts) >= 5:
            return 'simple'

        return None

    def _batch_simple(self, parts):
        """단순 측정값 일괄 입력"""
        item_no = parts[0]
        rows = int(parts[1])
        base = float(parts[2].replace('Ø', '').replace('ø', ''))
        upper_tol = float(parts[3])
        lower_tol = float(parts[4])
        ref = parts[5] if len(parts) > 5 else ""

        if lower_tol > 0:
            lower_tol = -lower_tol

        lower_calc = base + lower_tol
        upper_calc = base + upper_tol

        start_row = self.current_row
        self._unmerge_range(start_row, rows)

        for i in range(rows):
            row = self.current_row + i
            if i == 0:
                self.ws.cell(row, self.col_num, item_no)
            self.ws.cell(row, self.col_num + 1, base)
            self.ws.cell(row, self.col_num + 2, upper_tol)
            self.ws.cell(row, self.col_num + 3, lower_tol)
            self.ws.cell(row, self.col_num + 4, lower_calc)
            self.ws.cell(row, self.col_num + 5, upper_calc)
            self.ws.cell(row, self.col_num + 6, ref)

        if rows > 1:
            self.ws.merge_cells(start_row=start_row, start_column=self.col_num,
                              end_row=start_row + rows - 1, end_column=self.col_num)
            self.ws.cell(start_row, self.col_num).alignment = Alignment(horizontal='center', vertical='center')

        self.current_row += rows
        return f"[단순] 항목 {item_no}: {rows}개 행"

    def _batch_position(self, parts):
        """위치도 값 일괄 입력"""
        item_no = parts[0]
        rows = int(parts[1])
        base_str = parts[2]
        base = float(base_str.replace('Ø', '').replace('ø', ''))
        upper_tol = float(parts[3])
        lower_tol = float(parts[4])
        ref = parts[5] if len(parts) > 5 else ""

        if lower_tol > 0:
            lower_tol = -lower_tol

        lower_calc = base + lower_tol
        upper_calc = base + upper_tol

        start_row = self.current_row
        self._unmerge_range(start_row, rows)

        for i in range(rows):
            row = self.current_row + i
            if i == 0:
                self.ws.cell(row, self.col_num, item_no)
            self.ws.cell(row, self.col_num + 1, base_str)
            self.ws.cell(row, self.col_num + 2, upper_tol)
            self.ws.cell(row, self.col_num + 3, lower_tol)
            self.ws.cell(row, self.col_num + 4, lower_calc)
            self.ws.cell(row, self.col_num + 5, upper_calc)
            self.ws.cell(row, self.col_num + 6, ref)

        if rows > 1:
            self.ws.merge_cells(start_row=start_row, start_column=self.col_num,
                              end_row=start_row + rows - 1, end_column=self.col_num)
            self.ws.cell(start_row, self.col_num).alignment = Alignment(horizontal='center', vertical='center')

        self.current_row += rows
        return f"[위치도] 항목 {item_no}: {rows}개 행"

    def _batch_reference(self, parts):
        """참고 값 일괄 입력"""
        item_no = parts[0]
        rows = int(parts[1])
        base_str_with_paren = parts[2].strip()
        base_str_calc = parts[2].replace('(', '').replace(')', '').strip()

        has_tolerances = len(parts) >= 5 and parts[3] and parts[4]

        start_row = self.current_row
        self._unmerge_range(start_row, rows)

        for i in range(rows):
            row = self.current_row + i

            if i == 0:
                self.ws.cell(row, self.col_num, item_no)

            self.ws.cell(row, self.col_num + 1, base_str_with_paren)

            if has_tolerances:
                try:
                    base = float(base_str_calc)
                    upper_tol = float(parts[3])
                    lower_tol = float(parts[4])

                    if lower_tol > 0:
                        lower_tol = -lower_tol

                    lower_calc = base + lower_tol
                    upper_calc = base + upper_tol

                    self.ws.cell(row, self.col_num + 2, upper_tol)
                    self.ws.cell(row, self.col_num + 3, lower_tol)
                    self.ws.cell(row, self.col_num + 4, lower_calc)
                    self.ws.cell(row, self.col_num + 5, upper_calc)
                except ValueError:
                    for j in range(2, 6):
                        self.ws.cell(row, self.col_num + j, '-')
            else:
                for j in range(2, 6):
                    self.ws.cell(row, self.col_num + j, '-')

            if has_tolerances and len(parts) > 5:
                ref = parts[5]
            elif not has_tolerances and len(parts) > 3:
                ref = parts[3]
            else:
                ref = "참고"

            self.ws.cell(row, self.col_num + 6, ref)

        if rows > 1:
            self.ws.merge_cells(start_row=start_row, start_column=self.col_num,
                              end_row=start_row + rows - 1, end_column=self.col_num)
            self.ws.cell(start_row, self.col_num).alignment = Alignment(horizontal='center', vertical='center')

        self.current_row += rows
        return f"[참고] 항목 {item_no}: {rows}개 행"

    def _batch_mmc(self, parts):
        """MMC 공차 일괄 입력"""
        item_no = parts[0]
        num_sets = int(parts[1])
        mmc_str = parts[2].lower().replace('mmc', '').replace('(', '').replace(')', '').replace('m', '').strip()
        mmc_tol = float(mmc_str)
        max_val = parts[3] if len(parts) > 3 and parts[3] else ""
        ref = parts[4] if len(parts) > 4 else ""

        total_rows = num_sets * 3
        start_row = self.current_row

        self._unmerge_range(start_row, total_rows)

        for set_idx in range(num_sets):
            base_row = self.current_row + (set_idx * 3)

            # 1행: MMC 기준값 행
            if set_idx == 0:
                self.ws.cell(base_row, self.col_num, item_no)
            self.ws.cell(base_row, self.col_num + 1, f"{mmc_tol}ⓜ")
            self.ws.cell(base_row, self.col_num + 2, 0)
            self.ws.cell(base_row, self.col_num + 3, mmc_tol)
            self.ws.cell(base_row, self.col_num + 4, 0)
            self.ws.cell(base_row, self.col_num + 5, mmc_tol)
            self.ws.cell(base_row, self.col_num + 6, ref)

            # 2행: MAX값 행
            if max_val:
                try:
                    self.ws.cell(base_row + 1, self.col_num + 1, float(max_val))
                except:
                    self.ws.cell(base_row + 1, self.col_num + 1, max_val)
            for i in range(2, 6):
                self.ws.cell(base_row + 1, self.col_num + i, '-')
            self.ws.cell(base_row + 1, self.col_num + 6, "MMC 공차")

            # 3행: 측정값 입력 빈 칸
            for i in range(2, 6):
                self.ws.cell(base_row + 2, self.col_num + i, '-')
            self.ws.cell(base_row + 2, self.col_num + 6, ref)

        self.ws.merge_cells(start_row=start_row, start_column=self.col_num,
                          end_row=start_row + total_rows - 1, end_column=self.col_num)
        self.ws.cell(start_row, self.col_num).alignment = Alignment(horizontal='center', vertical='center')

        self.current_row += total_rows
        return f"[MMC] 항목 {item_no}: {num_sets}세트 ({total_rows}개 행)"

    def process_batch(self, lines):
        """일괄 입력 처리"""
        results = []
        count = 0

        for line in lines:
            parts = [p.strip() for p in line.split(',')]
            if len(parts) < 2:
                results.append(f"⚠ 형식 오류 (최소 2개 항목 필요): {line}")
                continue

            try:
                data_type = self._detect_data_type(parts, line)

                if data_type == 'simple':
                    msg = self._batch_simple(parts)
                elif data_type == 'position':
                    msg = self._batch_position(parts)
                elif data_type == 'reference':
                    msg = self._batch_reference(parts)
                elif data_type == 'mmc':
                    msg = self._batch_mmc(parts)
                else:
                    results.append(f"⚠ 유형을 감지할 수 없음: {line}")
                    continue

                results.append(f"✓ {msg}")
                count += 1

            except Exception as e:
                results.append(f"⚠ 오류: {line} - {str(e)}")

        return results, count

# Streamlit UI
st.set_page_config(
    page_title="엑셀 측정 데이터 입력 시스템",
    page_icon="📊",
    layout="wide"
)

st.title("📊 엑셀 측정 데이터 입력 시스템")
st.markdown("---")

# 사이드바 - 파일 업로드 및 설정
with st.sidebar:
    st.header("⚙️ 설정")

    uploaded_file = st.file_uploader(
        "엑셀 파일 업로드",
        type=['xlsx', 'xlsm'],
        help="측정 데이터를 입력할 엑셀 파일을 선택하세요"
    )

    if uploaded_file:
        # 파일 로드
        wb = load_workbook(io.BytesIO(uploaded_file.read()))

        # 시트 선택
        sheet_names = wb.sheetnames
        selected_sheet = st.selectbox(
            "시트 선택",
            sheet_names,
            index=sheet_names.index(wb.active.title) if wb.active.title in sheet_names else 0
        )

        # 시작 위치 설정
        col1, col2 = st.columns(2)
        with col1:
            start_col = st.text_input("시작 열", value="F")
        with col2:
            start_row = st.number_input("시작 행", min_value=1, value=5, step=1)

        st.markdown("---")
        st.info(f"📍 입력 위치: **{start_col}{start_row}**")

# 메인 영역
if uploaded_file:
    st.header("📝 데이터 입력")

    # 입력 형식 가이드
    with st.expander("📖 입력 형식 가이드", expanded=False):
        st.markdown("""
        ### 1️⃣ 단순 측정값
        **형식:** `항목번호, 행개수, 기준값, 상한공차, 하한공차`
        **예시:** `51, 1, 7.0, 0.15, 0.15`
        **결과:** 7.0, 0.15, -0.15, 6.85, 7.15

        ### 2️⃣ 위치도 값 (기준값에 Ø 포함)
        **형식:** `항목번호, 행개수, 기준값, 상한공차, 하한공차, [REF]`
        **예시:** `55, 4, Ø4.25, 0.15, 0.15`

        ### 3️⃣ 참고 값 (기준값을 괄호로 감싸기)
        **형식:** `항목번호, 행개수, (기준값), [상한공차], [하한공차], [REF]`
        **예시:** `60, 1, (1.2)` 또는 `61, 3, (7.0), 0.15, 0.15, 참고`

        ### 4️⃣ MMC 공차 (MMC 값에 'm' 붙이기)
        **형식:** `항목번호, 세트개수, MMC공차, [MAX값]`
        **예시:** `70, 10, 0.2m` 또는 `70, 10, 0.2m, 0.5`

        ---

        **💡 팁:**
        - 하한공차는 양수로 입력해도 자동으로 음수 변환
        - 여러 행이 있는 항목은 자동으로 번호 셀 병합
        - 한 줄에 한 항목씩 입력
        """)

    # 일괄 입력
    batch_input = st.text_area(
        "데이터 입력 (한 줄에 하나씩)",
        height=300,
        placeholder="""51, 1, 7.0, 0.15, 0.15
52, 3, 10.5, 0.2, 0.1
55, 4, Ø4.25, 0.15, 0.15
60, 1, (1.2)
61, 3, (7.0), 0.15, 0.15, 참고
70, 10, 0.2m, 0.5"""
    )

    col1, col2, col3 = st.columns([1, 1, 2])

    with col1:
        process_btn = st.button("✨ 데이터 처리", type="primary", use_container_width=True)

    with col2:
        clear_btn = st.button("🗑️ 입력 초기화", use_container_width=True)

    if clear_btn:
        st.rerun()

    # 데이터 처리
    if process_btn and batch_input:
        lines = [line.strip() for line in batch_input.split('\n') if line.strip()]

        if lines:
            with st.spinner('데이터 처리 중...'):
                processor = StreamlitExcelInput(wb, selected_sheet, start_col, start_row)
                results, count = processor.process_batch(lines)

                st.success(f"✅ 처리 완료! 총 {count}개 항목이 입력되었습니다.")

                # 결과 표시
                with st.expander("📋 처리 결과", expanded=True):
                    for result in results:
                        if "✓" in result:
                            st.success(result)
                        elif "⚠" in result:
                            st.warning(result)

                # 다운로드 버튼
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)

                st.download_button(
                    label="💾 결과 파일 다운로드",
                    data=output,
                    file_name=f"입력완료_{uploaded_file.name}",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
        else:
            st.error("❌ 입력된 데이터가 없습니다.")

    elif process_btn:
        st.warning("⚠️ 데이터를 입력해주세요.")

else:
    # 파일이 업로드되지 않았을 때
    st.info("👈 왼쪽 사이드바에서 엑셀 파일을 업로드하세요")

    st.markdown("### 📚 사용 방법")
    st.markdown("""
    1. 왼쪽 사이드바에서 엑셀 파일 업로드
    2. 시트 선택 및 시작 위치 설정
    3. 데이터 입력란에 측정 데이터 입력
    4. '데이터 처리' 버튼 클릭
    5. 처리 완료 후 '결과 파일 다운로드' 버튼으로 저장
    """)

# 푸터
st.markdown("---")
st.caption("📊 엑셀 측정 데이터 입력 시스템 v1.0 | Powered by Streamlit")
